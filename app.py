from flask import Flask, render_template, request, redirect, url_for, flash, send_file, send_from_directory, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from PIL import Image
import os
import json
import uuid
from datetime import datetime, date
from threading import Thread
from models import CertificateTemplate
from certificate_generator import generate_enhanced_certificate
from datetime import datetime

# Import models and other modules
from models import db, User, Student, Event, Certificate, EventParticipant, CertificateTemplate, EventType
from certificate_generator import generate_certificate_pdf, generate_bulk_certificates
from template_config import create_default_templates
from email_sender import send_email  # Import your working email sender
from email_sender import send_certificate_email_flask

# Initialize Flask app
app = Flask(__name__)

# App Configuration
app.config['SECRET_KEY'] = 'your-super-secret-key-change-in-production-2025'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///certificates.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Upload configurations
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['CERTIFICATE_FOLDER'] = 'certificates'
app.config['LOGO_UPLOAD_FOLDER'] = 'static/uploads/logos'
app.config['SIGNATURE_UPLOAD_FOLDER'] = 'static/uploads/signatures'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Initialize extensions
db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'

@login_manager.user_loader
def load_user(user_id):
    try:
        return User.query.get(int(user_id))
    except:
        return None

# Create directories
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['CERTIFICATE_FOLDER'], exist_ok=True)
os.makedirs(app.config['LOGO_UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['SIGNATURE_UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('static/templates', exist_ok=True)
os.makedirs('static/css', exist_ok=True)
os.makedirs('static/js', exist_ok=True)

# Helper Functions
def allowed_file(filename):
    """Check if file extension is allowed for images"""
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'svg'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_uploaded_image(file, upload_folder, prefix="img"):
    """Save uploaded image with unique filename and optimization"""
    if file and allowed_file(file.filename):
        # Generate unique filename
        file_ext = file.filename.rsplit('.', 1)[1].lower()
        unique_filename = f"{prefix}_{uuid.uuid4().hex}.{file_ext}"
        
        filepath = os.path.join(upload_folder, unique_filename)
        
        try:
            if file_ext.lower() in ['jpg', 'jpeg', 'png']:
                # Open and optimize image
                image = Image.open(file)
                
                # Convert RGBA to RGB if necessary
                if image.mode == 'RGBA':
                    background = Image.new('RGB', image.size, (255, 255, 255))
                    background.paste(image, mask=image.split()[-1])
                    image = background
                
                # Resize if larger than 1200x800
                if image.width > 1200 or image.height > 800:
                    image.thumbnail((1200, 800), Image.Resampling.LANCZOS)
                
                # Save with optimization
                image.save(filepath, optimize=True, quality=85)
            else:
                # For SVG and other formats, save directly
                file.save(filepath)
            
            return unique_filename
        except Exception as e:
            print(f"Error saving image: {e}")
            return None
    return None
    
def delete_image_file(filename, folder):
    """Delete image file from filesystem"""
    try:
        file_path = os.path.join(folder, filename)
        if os.path.exists(file_path):
            os.remove(file_path)
            return True
    except Exception as e:
        print(f"Error deleting file: {e}")
    return False

# Initialize database on first request
_initialization_done = False

@app.before_request
def initialize_database():
    """Initialize database on first request"""
    global _initialization_done
    if not _initialization_done:
        try:
            db.create_all()
            
            # Create admin user if not exists
            admin = User.query.filter_by(username='admin').first()
            if not admin:
                hashed_password = generate_password_hash('admin123', method='pbkdf2:sha256')
                admin = User(
                    username='admin',
                    email='admin@certmanager.com',
                    password=hashed_password,
                    is_admin=True
                )
                db.session.add(admin)
                db.session.commit()
                print("Admin user created: admin/admin123")
            
            # Create default templates
            create_default_templates()
            _initialization_done = True
            
        except Exception as e:
            print(f"Error initializing database: {e}")

# Context processors for templates
@app.context_processor
def inject_current_year():
    return {'current_year': datetime.now().year}

@app.context_processor
def inject_event_types():
    return {'EventType': EventType}

# ROUTES

@app.route('/')
def index():
    """Enhanced home page with statistics and recent events"""
    # Initialize stats
    students_count = 0
    events_count = 0
    certificates_count = 0
    emails_sent = 0
    recent_events = []
    
    # Get statistics if user is authenticated
    if current_user.is_authenticated:
        try:
            students_count = Student.query.count()
            events_count = Event.query.count()
            certificates_count = Certificate.query.count()
            
            # Get recent events (last 5)
            recent_events = Event.query.order_by(Event.created_at.desc()).limit(5).all()
            
            # Estimate emails sent (certificates * events as rough estimate)
            emails_sent = certificates_count
            
        except Exception as e:
            print(f"Error getting statistics: {e}")
    
    return render_template('index.html',
                         students_count=students_count,
                         events_count=events_count,
                         certificates_count=certificates_count,
                         emails_sent=emails_sent,
                         recent_events=recent_events)

@app.route('/register', methods=['GET', 'POST'])
def register():
    """User registration"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        
        # Validation
        if not username or not email or not password:
            flash('All fields are required', 'error')
            return render_template('register.html')
        
        if len(username) < 3:
            flash('Username must be at least 3 characters long', 'error')
            return render_template('register.html')
        
        if len(password) < 6:
            flash('Password must be at least 6 characters long', 'error')
            return render_template('register.html')
        
        # Check if user already exists
        if User.query.filter_by(username=username).first():
            flash('Username already exists. Please choose a different one.', 'error')
            return render_template('register.html')
        
        if User.query.filter_by(email=email).first():
            flash('Email already registered. Please use a different email.', 'error')
            return render_template('register.html')
        
        # Create new user
        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
        new_user = User(
            username=username, 
            email=email, 
            password=hashed_password
        )
        
        try:
            db.session.add(new_user)
            db.session.commit()
            flash('Registration successful! You can now login.', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            flash(f'Registration failed: {str(e)}', 'error')
    
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    """User login"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        if not username or not password:
            flash('Username and password are required', 'error')
            return render_template('login.html')
        
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password, password):
            login_user(user)
            flash(f'Welcome back, {user.username}!', 'success')
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    """User logout"""
    username = current_user.username
    logout_user()
    flash(f'Goodbye {username}! You have been logged out successfully.', 'info')
    return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard with statistics"""
    students_count = Student.query.count()
    events_count = Event.query.count()
    certificates_count = Certificate.query.count()
    
    # Event type statistics
    event_stats = {}
    for event_type in EventType:
        count = Event.query.filter_by(event_type=event_type).count()
        event_stats[event_type.value] = count
    
    # Recent activities
    recent_events = Event.query.order_by(Event.created_at.desc()).limit(5).all()
    recent_certificates = Certificate.query.order_by(Certificate.issued_date.desc()).limit(5).all()
    
    return render_template('dashboard.html', 
                         students_count=students_count,
                         events_count=events_count,
                         certificates_count=certificates_count,
                         event_stats=event_stats,
                         recent_events=recent_events,
                         recent_certificates=recent_certificates)


@app.route('/events', methods=['GET', 'POST'])
def events():
    if request.method == 'POST':
        title = request.form['title']
        event_type = request.form['event_type']
        organizer = request.form['organizer']
        location = request.form['location']
        start_date = datetime.strptime(request.form['start_date'], "%Y-%m-%d").date()
        end_date = datetime.strptime(request.form['end_date'], "%Y-%m-%d").date()
        year = int(request.form['year'])
        description = request.form.get('description')

        # Upload background image
        background_path = None
        if 'background' in request.files:
            bg_file = request.files['background']
            if bg_file and allowed_file(bg_file.filename):
                filename = str(uuid.uuid4()) + os.path.splitext(bg_file.filename)[1]
                bg_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                bg_file.save(bg_path)
                background_path = filename

                logo_path = None
        if 'logo' in request.files:
            logo_file = request.files['logo']
            print(f"Logo file: {logo_file.filename}")
            if logo_file and logo_file.filename != '' and allowed_file(logo_file.filename):
                filename = f"logo_{uuid.uuid4()}{os.path.splitext(logo_file.filename)[1]}"
                save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                logo_file.save(save_path)
                logo_path = filename
                print(f"Logo saved as: {logo_path}")
        
        # Handle signature upload
        signature_path = None
        if 'signature' in request.files:
            sig_file = request.files['signature']
            print(f"Signature file: {sig_file.filename}")
            if sig_file and sig_file.filename != '' and allowed_file(sig_file.filename):
                filename = f"sig_{uuid.uuid4()}{os.path.splitext(sig_file.filename)[1]}"
                save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                sig_file.save(save_path)
                signature_path = filename
                print(f"Signature saved as: {signature_path}")

        # Save event
        event = Event(
        title=request.form['title'],
        event_type=EventType[request.form['event_type']],
        organizer=request.form['organizer'],
        location=request.form['location'],
        teacher_id=current_user.id,
        start_date=start_date,
        end_date=end_date,
        date=start_date,   # <---- THIS LINE PREVENTS THE ERROR
        year=int(request.form['year']),
        description=request.form.get('description'),
        background_path=background_path,
        logo_path=logo_path,
        signature_path=signature_path,
        created_by=current_user.id
    )

        db.session.add(event)
        db.session.commit()
        flash("Event created!", "success")
        return redirect(url_for('events'))

    events = Event.query.order_by(Event.start_date.desc()).all()
    return render_template('events.html', events=events, current_year=datetime.now().year)

@app.route('/preview_certificate/<int:event_id>/<int:student_id>')
def preview_certificate_html(event_id, student_id):
    # Logic to preview certificate
    return render_template('certificate_preview.html', 
                         event_id=event_id, 
                         student_id=student_id)

@app.route('/upload_students', methods=['GET', 'POST'])
@app.route('/upload_students/<int:event_id>', methods=['GET', 'POST'])
@login_required
def upload_students(event_id=None):
    """Upload students via Excel file with multi-event support"""
    selected_event = None
    participants = []
    
    if event_id:
        selected_event = Event.query.get_or_404(event_id)
        participants = EventParticipant.query.filter_by(event_id=event_id).all()
    
    if request.method == 'POST':
        event_id = request.form.get('event_id')
        if not event_id:
            flash('Please select an event', 'error')
            return redirect(request.url)
        
        event = Event.query.get_or_404(event_id)
        
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(request.url)
        
        # Save and process file
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            workbook = load_workbook(filepath)
            sheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).lower().strip().replace(' ', '_'))
                else:
                    headers.append('')
            
            students_added = 0
            students_registered = 0
            students_reused = 0
            errors = []
            
            # Process each row
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                try:
                    if not row or not any(row):
                        continue
                    
                    # Extract student data
                    student_data = {}
                    for i, value in enumerate(row):
                        if i < len(headers) and headers[i] and value:
                            student_data[headers[i]] = str(value).strip()
                    
                    # Get name and email (required fields)
                    name = (student_data.get('name') or 
                           student_data.get('student_name') or 
                           student_data.get('full_name') or 
                           (str(row[0]).strip() if row[0] else None))
                    
                    email = (student_data.get('email') or 
                            student_data.get('email_address') or 
                            (str(row[1]).strip() if len(row) > 1 and row[1] else None))
                    
                    if not name or not email:
                        errors.append(f'Row {row_num}: Missing name or email')
                        continue
                    
                    # Check if student already exists (by email)
                    student = Student.query.filter_by(email=email).first()
                    is_new_student = False
                    
                    if not student:
                        # Create new student
                        student = Student(
                            name=name,
                            email=email,
                            student_id=student_data.get('student_id'),
                            phone=student_data.get('phone'),
                            department=student_data.get('department'),
                            course=student_data.get('course')
                        )
                        db.session.add(student)
                        db.session.flush()  # Get student ID
                        students_added += 1
                        is_new_student = True
                    else:
                        # Student exists, will be reused for this event
                        students_reused += 1
                    
                    # Check if already registered for this event
                    existing_participation = EventParticipant.query.filter_by(
                        event_id=event.id, 
                        student_id=student.id
                    ).first()
                    
                    if not existing_participation:
                        # Register student for this event
                        participation = EventParticipant(
                            event_id=event.id,
                            student_id=student.id,
                            participation_type=student_data.get('participation_type', 'Participant'),
                            achievement_level=student_data.get('achievement_level'),
                            special_recognition=student_data.get('special_recognition')
                        )
                        db.session.add(participation)
                        students_registered += 1
                    else:
                        if not is_new_student:
                            students_reused -= 1  # Don't count as reused if already registered
                        
                except Exception as row_error:
                    errors.append(f'Row {row_num}: {str(row_error)}')
                    continue
            
            # Commit all changes
            db.session.commit()
            
            # Success message
            success_parts = []
            if students_added > 0:
                success_parts.append(f'{students_added} new students created')
            if students_reused > 0:
                success_parts.append(f'{students_reused} existing students reused')
            success_parts.append(f'{students_registered} students registered for "{event.title}"')
            
            flash(f'Success! {", ".join(success_parts)}', 'success')
            
            # Show errors if any
            if errors:
                for error in errors[:5]:  # Show max 5 errors
                    flash(error, 'warning')
                if len(errors) > 5:
                    flash(f'...and {len(errors) - 5} more errors', 'warning')
            
            # Cleanup
            os.remove(filepath)
            return redirect(url_for('upload_students', event_id=event.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error processing file: {str(e)}', 'error')
            if os.path.exists(filepath):
                os.remove(filepath)
    
    # Get all events for dropdown
    events = Event.query.order_by(Event.created_at.desc()).all()
    return render_template('upload_students.html', 
                         events=events, 
                         selected_event=selected_event,
                         participants=participants)

@app.route('/bulk_certificates', methods=['GET', 'POST'])
@login_required
def bulk_certificates():
    """Bulk certificate generation for events"""
    if request.method == 'POST':
        event_id = request.form.get('event_id')
        template_id = request.form.get('template_id')
        
        if not event_id:
            flash('Please select an event', 'error')
            return redirect(request.url)
        
        event = Event.query.get_or_404(event_id)
        participants = EventParticipant.query.filter_by(event_id=event_id).all()
        students = [p.student for p in participants]
        
        if not students:
            flash('No students registered for this event', 'warning')
            return redirect(request.url)
        
        try:
            # Generate certificates for all students
            pdf_paths = generate_bulk_certificates(event, students, app.config['CERTIFICATE_FOLDER'], template_id)
            
            successful_certs = 0
            failed_certs = 0
            
            for student, pdf_path in zip(students, pdf_paths):
                if pdf_path:
                    # Record certificate in database
                    existing_cert = Certificate.query.filter_by(
                        student_id=student.id, 
                        event_id=event_id
                    ).first()
                    
                    if not existing_cert:
                        certificate = Certificate(
                            student_id=student.id, 
                            event_id=event_id, 
                            template_id=template_id,
                            certificate_path=pdf_path
                        )
                        db.session.add(certificate)
                    successful_certs += 1
                else:
                    failed_certs += 1
            
            db.session.commit()
            
            # Success message
            if successful_certs > 0:
                flash(f'Successfully generated {successful_certs} certificates for {event.title} ({event.event_type.value})!', 'success')
            if failed_certs > 0:
                flash(f'{failed_certs} certificates failed to generate', 'warning')
                
        except Exception as e:
            db.session.rollback()
            flash(f'Error generating bulk certificates: {str(e)}', 'error')
    
    # Get data for template
    events = Event.query.order_by(Event.created_at.desc()).all()
    templates = CertificateTemplate.query.filter_by(is_active=True).all()
    
    # Add participant count to events
    for event in events:
        event.participant_count = len(event.participants)
    
    return render_template('bulk_certificates.html', events=events, templates=templates)

# EMAIL ROUTES

@app.route('/test_email')
@login_required
def test_email():
    """Test email functionality"""
    try:
        subject = "🧪 Test Email from CertManager"
        recipient = current_user.email
        body = f"""
Hello {current_user.username},

This is a test email to confirm your email configuration is working properly.

✅ Email system is functioning correctly!

Best regards,
CertManager System
"""
        
        if send_email(subject, recipient, body):
            flash(f'Test email sent successfully to {recipient}!', 'success')
        else:
            flash('Failed to send test email. Please check configuration.', 'error')
            
    except Exception as e:
        flash(f'Error sending test email: {str(e)}', 'error')
    
    return redirect(url_for('dashboard'))

@app.route('/send_certificate_email/<int:event_id>/<int:student_id>')
@login_required
def send_certificate_email(event_id, student_id):
    event = Event.query.get_or_404(event_id)
    student = Student.query.get_or_404(student_id)
    
    participation = EventParticipant.query.filter_by(event_id=event_id, student_id=student_id).first()
    if not participation:
        flash('Student is not registered for this event', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        pdf_path = generate_certificate_pdf(event, student, app.config['CERTIFICATE_FOLDER'])
        if pdf_path:
            with open(pdf_path, 'rb') as f:
                pdf_data = f.read()
            download_url = url_for('generate_pdf', event_id=event.id, student_id=student.id, _external=True)
            # Main change: use the new email sender function!
            ok = send_certificate_email_flask(student, event, pdf_data, download_url)
            if ok:
                # Optionally add to Certificate table, etc...
                flash(f'Certificate notification sent to {student.email}!', 'success')
            else:
                flash('Failed to send certificate email', 'error')
        else:
            flash('Error generating certificate', 'error')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(request.referrer or url_for('upload_students', event_id=event_id))


@app.route('/bulk_email_certificates', methods=['POST'])
@login_required
def bulk_email_certificates():
    """Send certificates to all students in an event via email"""
    event_id = request.form.get('event_id')
    
    if not event_id:
        flash('Please select an event', 'error')
        return redirect(url_for('bulk_certificates'))
    
    event = Event.query.get_or_404(event_id)
    participants = EventParticipant.query.filter_by(event_id=event_id).all()
    
    if not participants:
        flash('No students registered for this event', 'warning')
        return redirect(url_for('bulk_certificates'))
    
    successful_emails = 0
    failed_emails = 0
    
    for participation in participants:
        student = participation.student
        
        try:
            # Generate certificate
            pdf_path = generate_certificate_pdf(event, student, app.config['CERTIFICATE_FOLDER'])
            
            if pdf_path:
                subject = f"🏆 Your Certificate - {event.title}"
                
                body = f"""
Dear {student.name},

Your certificate for "{event.title}" is ready!

📋 Event Details:
• Event: {event.title} ({event.event_type.value})
• Date: {event.date.strftime('%B %d, %Y')}
• Organizer: {event.organizer}
• Location: {event.location}

Download Link: {url_for('generate_pdf', event_id=event_id, student_id=student.id, _external=True)}

Congratulations on your participation!

Best regards,
{event.organizer}
"""
                
                if send_email(subject, student.email, body):
                    successful_emails += 1
                    
                    # Record certificate
                    existing_cert = Certificate.query.filter_by(student_id=student.id, event_id=event_id).first()
                    if not existing_cert:
                        certificate = Certificate(
                            student_id=student.id,
                            event_id=event_id,
                            certificate_path=pdf_path
                        )
                        db.session.add(certificate)
                else:
                    failed_emails += 1
            else:
                failed_emails += 1
                
        except Exception as e:
            print(f"Error sending email to {student.email}: {e}")
            failed_emails += 1
    
    db.session.commit()
    
    # Show results
    if successful_emails > 0:
        flash(f'✅ Successfully sent {successful_emails} certificate notifications!', 'success')
    if failed_emails > 0:
        flash(f'❌ {failed_emails} emails failed to send', 'warning')
    
    return redirect(url_for('bulk_certificates'))

@app.route('/send_custom_email', methods=['POST'])
@login_required
def send_custom_email():
    """Send custom email to selected students"""
    recipient_emails = request.form.getlist('recipient_emails')
    subject = request.form.get('subject', '').strip()
    message = request.form.get('message', '').strip()
    
    if not recipient_emails or not subject or not message:
        flash('Please fill all fields and select recipients', 'error')
        return redirect(request.referrer)
    
    successful = 0
    failed = 0
    
    for email in recipient_emails:
        try:
            if send_email(subject, email, message):
                successful += 1
            else:
                failed += 1
        except Exception as e:
            print(f"Error sending to {email}: {e}")
            failed += 1
    
    if successful > 0:
        flash(f'✅ Sent {successful} emails successfully!', 'success')
    if failed > 0:
        flash(f'❌ {failed} emails failed to send', 'warning')
    
    return redirect(request.referrer)

# PDF GENERATION ROUTES

@app.route('/generate_pdf/<int:event_id>/<int:student_id>')
@app.route('/generate_pdf/<int:event_id>/<int:student_id>/<int:template_id>')
@login_required
def generate_pdf(event_id, student_id, template_id=None):
    """Generate and download individual certificate PDF"""
    event = Event.query.get_or_404(event_id)
    student = Student.query.get_or_404(student_id)

    pdf_path = generate_enhanced_certificate(
            event=event, 
            student=student, 
            certificate_folder=app.config['CERTIFICATE_FOLDER']
        )
    
    # Check if student is registered for this event
    participation = EventParticipant.query.filter_by(event_id=event_id, student_id=student_id).first()
    if not participation:
        flash('Student is not registered for this event', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        pdf_path = generate_certificate_pdf(event, student, app.config['CERTIFICATE_FOLDER'], template_id)
        
        if pdf_path:
            # Record certificate generation
            existing_cert = Certificate.query.filter_by(student_id=student_id, event_id=event_id).first()
            if not existing_cert:
                certificate = Certificate(
                    student_id=student_id, 
                    event_id=event_id, 
                    template_id=template_id,
                    certificate_path=pdf_path
                )
                db.session.add(certificate)
                db.session.commit()
            
            # Send file
            filename = f'certificate_{student.name.replace(" ", "_")}_{event.title.replace(" ", "_")}.pdf'
            return send_file(pdf_path, as_attachment=True, download_name=filename)
        else:
            flash('Error generating certificate. Please try again.', 'error')
            return redirect(url_for('dashboard'))
            
    except Exception as e:
        flash(f'Error generating certificate: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

# File serving routes
@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    """Serve uploaded logo and signature files"""
    if filename.startswith('logo_'):
        return send_from_directory(app.config['LOGO_UPLOAD_FOLDER'], filename)
    elif filename.startswith('signature_'):
        return send_from_directory(app.config['SIGNATURE_UPLOAD_FOLDER'], filename)
    else:
        return send_from_directory(app.config['LOGO_UPLOAD_FOLDER'], filename)

@app.route('/delete_event_image/<int:event_id>/<image_type>')
@login_required
def delete_event_image(event_id, image_type):
    """Delete event logo or signature"""
    event = Event.query.get_or_404(event_id)
    
    if image_type == 'logo' and event.logo_path:
        if delete_image_file(event.logo_path, app.config['LOGO_UPLOAD_FOLDER']):
            event.logo_path = None
            flash('Logo deleted successfully!', 'success')
        else:
            flash('Error deleting logo file', 'error')
            
    elif image_type == 'signature' and event.signature_path:
        if delete_image_file(event.signature_path, app.config['SIGNATURE_UPLOAD_FOLDER']):
            event.signature_path = None
            flash('Signature deleted successfully!', 'success')
        else:
            flash('Error deleting signature file', 'error')
    else:
        flash('Image not found or invalid image type', 'warning')
    
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Database error: {str(e)}', 'error')
    
    return redirect(url_for('events'))

# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('errors/500.html'), 500

@app.errorhandler(403)
def forbidden_error(error):
    return render_template('errors/403.html'), 403

@app.errorhandler(405)
def method_not_allowed_error(error):
    return render_template('errors/405.html'), 405

# Additional important page routes
@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/contact')
def contact():
    return render_template('contact.html')

@app.route('/privacy')
def privacy():
    return render_template('privacy.html', datetime=datetime)

@app.route('/terms')
def terms():
    return render_template('terms.html')

@app.route('/help')
def help_page():
    return render_template('help.html')

# Run the application
if __name__ == '__main__':
    with app.app_context():
        try:
            db.create_all()
            
            # Create admin user if not exists
            admin = User.query.filter_by(username='admin').first()
            if not admin:
                hashed_password = generate_password_hash('admin123', method='pbkdf2:sha256')
                admin = User(
                    username='admin',
                    email='admin@certmanager.com',
                    password=hashed_password,
                    is_admin=True
                )
                db.session.add(admin)
                db.session.commit()
                print("✅ Admin user created: admin/admin123")
            
            # Create default templates
            create_default_templates()
            print("✅ Database initialized successfully!")
            print("🚀 Starting Certificate Manager...")
            
        except Exception as e:
            print(f"❌ Error initializing database: {e}")
    
    # Run Flask app
    app.run(debug=True, port=5000)